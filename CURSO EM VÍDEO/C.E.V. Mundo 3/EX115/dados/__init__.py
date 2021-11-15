def criar_arquivo(nome="Padrão.xlsx"):
    """
    Testa se arquivo já existe.
    :param nome: nome do arquivo em questão.
    :return: cria o arquivo, caso não exista, e retorna o nome do mesmo.
    """
    from openpyxl import Workbook
    from openpyxl import load_workbook
    try:  # Testa se o arquivo já existe.
        ler = load_workbook(filename=nome)
    except:  # Caso a resposta seja não, cria o arquivo em questão.
        lista_completa = Workbook()
        lista_completa.save(filename=nome)
    return nome


def ler_dados(nome="Padrão.xlsx", colunaInicial: int = None, linhaInicial: int = None,
              linhaFinal: int = None, colunaFinal: int = None):
    """
    Lê o arquivo e retorna uma lista, com ou sem print.
    :param colunaFinal: coluna másima a ser lida
    :param linhaFinal: linha máxima a ser lida
    :param colunaInicial: coluna mínima a ser lida
    :param linhaInicial: linha mínima a ser lida
    :param nome:nome do arquivo
    :return:retorna os dados em uma tupla
    """
    from openpyxl import load_workbook
    try:
        arquivo = load_workbook(filename=nome)
        pasta_do_arquivo = arquivo.active
        lista_completa = []
        for pessoa in pasta_do_arquivo.iter_cols(min_row=linhaInicial, max_row=linhaFinal,
                                                 min_col=colunaInicial, max_col=colunaFinal,
                                                 values_only=True):
            lista_completa.append(pessoa)
        return lista_completa
    except:
        print("Não é possível ler o arquivo, tente novamente.")


def adicionar_dado(pasta="Padrão.xlsx", dado="", coluna="", linha=1):
    from openpyxl import load_workbook
    arquivo = load_workbook(pasta)
    planilha = arquivo.active
    celula = coluna + str(linha)
    planilha[celula] = dado
    arquivo.save(filename=pasta)
